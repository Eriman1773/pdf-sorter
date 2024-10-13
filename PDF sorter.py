import re
from fuzzywuzzy import fuzz
import os
import functools
import datetime
import time
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import shutil

# Export
# - complete_match(name, text)
# - find_name_in_string(name, text)
# - move_file(src_file_path, dest_dir)


def write_error_to_excel(file_path, error_data, column_names):
    """
    This function handles the creation of the Excel file and writing errors to it.
    :param file_path: Path to the Excel file where errors are logged.
    :param error_data: A list of error details to be written to the Excel file.
    """
    while True:
        try:
            # Check if the Excel file already exists
            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
                sheet = workbook.active
            else:
                # Create a new workbook and set headers if the file doesn't exist
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(column_names)
            break
        except (PermissionError, InvalidFileException):
            print(f"Excel file '{file_path}' is locked or in use. Retrying in 3 seconds...")
            time.sleep(3)

    # Add the new error data to the sheet
    sheet.append(error_data)

    # Try saving the workbook, handling potential permission issues
    while True:
        try:
            workbook.save(file_path)
            break
        except PermissionError:
            print(f"Cannot save Excel file '{file_path}' as it is locked. Retrying in 3 seconds...")
            time.sleep(3)


def debug_decorator(func, record_logs=True, record_errors=True, display_errors_in_full=True, display_errors_short=False,
                    display_info=True):
    log_dir = 'errors'  # Directory to store error logs
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)  # Create the directory if it doesn't exist
    excel_file = os.path.join(log_dir, 'error_log.xlsx')  # Path to the Excel file for logging errors

    @functools.wraps(func)
    def wrapper_debug(*args, **kwargs):
        try:
            # Execute the function
            result = func(*args, **kwargs)

            # Display debugging information if enabled
            if display_info:
                print(f"--- DEBUGGING FUNCTION '{func.__name__}' ---")
                print(f"Arguments: {args}")
                print(f"Keyword Arguments: {kwargs}")
                print(f"Result: {result}")
                print(f"---------------------------\n")
            return result

        except Exception as e:
            error_traceback = traceback.format_exc()

            # Log the error to Excel if `record_errors` is enabled
            if record_errors:
                error_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                error_data = [
                    error_time,  # Timestamp of the error
                    func.__name__,  # Function name
                    str(args),  # Function arguments
                    str(kwargs),  # Keyword arguments
                    type(e).__name__,  # Error type
                    str(e),  # Error message
                    error_traceback  # Full traceback
                ]

                # Call the helper function to write the error to the Excel file
                column_names = ["Timestamp", "Function Name", "Arguments", "Keyword Arguments", "Error Type", "Error Message","Traceback"]
                write_error_to_excel(excel_file, error_data, column_names)

                print(f"Error in function '{func.__name__}' logged to {excel_file}")

            # Display the error message in the console based on the settings
            if display_errors_short:
                print(f"Error in function '{func.__name__}'")
            elif display_errors_in_full:
                print(f"--- ERROR in function '{func.__name__}' ---")
                print(f"Arguments: {args}")
                print(f"Keyword Arguments: {kwargs}")
                print(f"Error Type: {type(e).__name__}")
                print(f"Error Message: {e}")
                print(f"Traceback: {error_traceback}")
                print(f"---------------------------\n")

    return wrapper_debug

def normalize_string(s):
    """Helper function to normalize strings: lowercases, removes extra spaces and common separators."""
    # Convert to lowercase
    s = s.lower()
    # Remove common separators and multiple spaces
    s = re.sub(r'[\s\-_,.]+', ' ', s).strip()
    return s

@debug_decorator
def find_name_in_string(name, text, threshold=80):
    """
    Function to search for a name in a given text, taking into account variations in capitalization,
    separators, and slight differences.

    Parameters:
    - name (str): The name to search for.
    - text (str): The string containing the text where the name might be present.
    - threshold (int): The minimum fuzzy match score (0-100) to consider a name as matched.

    Returns:
    - bool: True if a sufficiently close match is found, False otherwise.
    - int: The fuzzy match score (0-100) if a match is found, 0 otherwise.
    """
    # Normalize both the name and the text
    normalized_name = normalize_string(name)
    normalized_text = normalize_string(text)

    # Split the text into potential words/phrases
    words = normalized_text.split()

    # Try to match the normalized name against all possible substrings in the text
    for i in range(len(words)):
        for j in range(i + 1, len(words) + 1):
            substring = ' '.join(words[i:j])
            match_score = fuzz.ratio(normalized_name, substring)
            if match_score >= threshold:
                return True, match_score

    return False, 0

@debug_decorator
def complete_match(name, text):
    """
    Debug function to test the find_name_in_string function with a complete match.
    """
    normalized_name = normalize_string(name)
    normalized_text = normalize_string(text)
    if normalized_name in normalized_text:
        return True, "Identical"



@debug_decorator
def move_file(src_file_path, dest_dir, logs_bool=True, log_dir = 'errors', logs_path = 'info_log.xlsx' ):
    file_name, file_extension = os.path.splitext(os.path.basename(src_file_path))
    dest_file_path = os.path.join(dest_dir, file_name + file_extension)

    # Если файл уже существует в целевой директории
    if os.path.exists(dest_file_path):
        counter = 2  # начинаем счет с 2 (file (2).pdf)
        while True:
            new_file_name = f"{file_name} ({counter}){file_extension}"
            new_dest_file_path = os.path.join(dest_dir, new_file_name)
            if not os.path.exists(new_dest_file_path):
                dest_file_path = new_dest_file_path
                break
            counter += 1

    # Перемещаем файл
    shutil.move(src_file_path, dest_file_path)
    print(f"{src_file_path} moved to: {dest_file_path}")


    if logs_bool:
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)  # Create the directory if it doesn't exist
        info_file = os.path.join(log_dir, logs_path)  # Path to the Excel file for logging errors

        error_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        error_data = [
            error_time,  # Timestamp of the error
            src_file_path,  # Function name
            "> moved to >",
            dest_file_path,
        ]

        column_names = ["log_time",  "src_file_path",  "> moved to >", "dest_file_path"]
        write_error_to_excel(info_file, error_data, column_names)

    return f"{src_file_path} moved to: {dest_file_path}"



if __name__ == '__main__':
    move_file('52343456.py', 'errors' )
    input("End")

    test_cases = [
        ("John Doe", "AG123123 01.04.20024 John Doe. str Marsweg 62a"),
        ("Jöht Doe", "AG123123 01.04.20024 John -  Doe. str Marsweg 62a"),
        ("Jihn Doe", "AG123123 01.04.20024 John Doe. str Marsweg 62a"),
        ("john doe", "AG123123 01.04.20024 John Doe. str Marsweg 62a"),

        ("00str Marsweg 62a", "AG123123 01.04.20024 John Doe. str Marsweg 62a"),
        ("00John Doe", "The client's name is John Doe.")
    ]
    for name, text in test_cases:
        complete_match(name, text)
    for name, text in test_cases:
        find_name_in_string(name, text)
